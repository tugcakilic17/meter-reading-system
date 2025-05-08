from flask import Flask, request, jsonify
from inference_sdk import InferenceHTTPClient
from PIL import Image
import pandas as pd
from io import BytesIO
import base64
import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
import os
import openpyxl
import easyocr  


app = Flask(__name__)


CLIENT_FIRST_MODEL = InferenceHTTPClient(
    api_url="https://detect.roboflow.com",
    api_key="k0i9ZNwIPkTLqEEAtcIU"
)


EXCEL_FILE = "sayaç_verileri.xlsx"


def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Fatura Verileri"
        
        ws.append(["Sayaç Numarası", "Sayaç Değeri"])
        wb.save(EXCEL_FILE)
        print(f"{EXCEL_FILE} oluşturuldu.")
    else:
        print(f"{EXCEL_FILE} zaten mevcut.")


def write_to_excel(sayac_numarasi, sayac_degeri):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([sayac_numarasi, sayac_degeri])  
    wb.save(EXCEL_FILE)
    print(f"Veriler Excel'e yazıldı: {sayac_numarasi}, {sayac_degeri}")

@app.route('/predict', methods=['POST'])
def predict():
    try:
        
        data = request.get_json()

        
        if 'image' not in data or not isinstance(data['image'], str):
            return jsonify({"error": "Invalid input format. 'image' key missing or not a string."}), 400
        image_data = data['image']
        image = Image.open(BytesIO(base64.b64decode(image_data)))

        
        open_cv_image = np.array(image)
        open_cv_image = cv2.cvtColor(open_cv_image, cv2.COLOR_RGB2BGR)

        
        height, width, _ = open_cv_image.shape
        cropped_first_half = open_cv_image[:height // 2, :]  

        
        _, encoded_first_half = cv2.imencode('.jpg', cropped_first_half)
        base64_first_half = base64.b64encode(encoded_first_half).decode('utf-8')
        roboflow_results = CLIENT_FIRST_MODEL.infer(base64_first_half, model_id="digits-tblxu/1")

        
        detections = roboflow_results['predictions']
        confidence_threshold = 0.05

        
        min_h, max_h = 20, 100  
        min_w, max_w = 2, 60   

        
        sorted_detections_first = sorted(
            [d for d in detections if d['confidence'] > confidence_threshold and d['class'].isdigit() and
             min_h < d['height'] < max_h and min_w < d['width'] < max_w],
            key=lambda d: d['x']  
        )

        
        for detection in sorted_detections_first:
            x1, y1 = int(detection['x'] - detection['width'] / 2), int(detection['y'] - detection['height'] / 2)
            x2, y2 = int(detection['x'] + detection['width'] / 2), int(detection['y'] + detection['height'] / 2)
            cls = detection['class']

            
            cv2.rectangle(cropped_first_half, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.putText(cropped_first_half, cls, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

        
        detected_numbers_first = "".join([d["class"] for d in sorted_detections_first])
        print(f"First Model Detected Numbers: {detected_numbers_first}")

        
        #boxed_image_path = "boxed_image.jpg"
        #cv2.imwrite(boxed_image_path, cropped_first_half)

        
        cropped_second_half = open_cv_image[int(height * 0.75):, :]  

        
        reader = easyocr.Reader(['en', 'tr'])  
        ocr_results = reader.readtext(cropped_second_half)
        detected_text_second = " ".join([res[1] for res in ocr_results]).strip()
        print(f"Second Model Detected Text (EasyOCR): {detected_text_second}")

        
        #boxed_image2_path = "boxed_image2.jpg"
        #cv2.imwrite(boxed_image2_path, cropped_second_half)

        
        return jsonify({
            "message": "Prediction completed.",
            "first_model_numbers": detected_numbers_first,
            "second_model_text": detected_text_second,
            #"boxed_image_path": boxed_image_path,
            #"boxed_image2_path": boxed_image2_path
        })

    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500



BASE_DIR = os.path.dirname(os.path.abspath(__file__))


EXCEL_FILE = os.path.join(BASE_DIR, "elektrik_fatura.xlsx")

@app.route('/update_excel', methods=['POST'])
def update_excel():
    try:
        data = request.get_json()
        print(f"Gelen veri: {data}")
        
        data = request.get_json()
        block = data.get('block', '').strip().upper()
        office_no = data.get('office_no', '').strip()
        month = data.get('month', '').strip().upper()
        kwh_value = data.get('kwh_value', 0.0)

        
        if not os.path.exists(EXCEL_FILE):
            return jsonify({"error": "Excel dosyası bulunamadı."}), 404

        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb["ELEKTRİK FATURASI"]

        
        row_index = None
        for row in range(2, sheet.max_row + 1):  # Satırları dolaş
            if (sheet[f"P{row}"].value == block) and (str(sheet[f"B{row}"].value) == office_no):
                row_index = row
                break

        if row_index is None:
            return jsonify({"error": "Blok ve Sayaç Numarasına uygun satır bulunamadı."}), 400

        
        column_index = None
        for col in range(4, sheet.max_column + 1): 
            if sheet.cell(row=1, column=col).value.strip().upper() == month:
                column_index = col
                break

        if column_index is None:
            return jsonify({"error": f"'{month}' ayına uygun bir sütun bulunamadı."}), 400

        
        sheet.cell(row=row_index, column=column_index, value=kwh_value)

        
        wb.save(EXCEL_FILE)

        return jsonify({"message": f"{month} sütununa {kwh_value} kWh değeri başarıyla yazıldı."}), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
